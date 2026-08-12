"""회사 양식에 맞는 엑셀 파일을 생성한다."""

import os
from datetime import date
from PIL import Image

from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill, numbers,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.drawing.image import Image as XlImage

from expense_processor import ExpenseRow, ProcessResult

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

FONT_TITLE = Font(name="맑은 고딕", size=16, bold=True, underline="single")
FONT_LABEL = Font(name="맑은 고딕", size=11, bold=True)
FONT_HEADER = Font(name="맑은 고딕", size=11, bold=True)
FONT_HEADER_SMALL = Font(name="맑은 고딕", size=9, bold=True)
FONT_DATA = Font(name="맑은 고딕", size=10)
FONT_NOTE = Font(name="맑은 고딕", size=10)
FONT_NOTE_BOLD = Font(name="맑은 고딕", size=10, bold=True, underline="single")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_CENTER_SHRINK = Alignment(horizontal="center", vertical="center", shrink_to_fit=True)

WON_FORMAT = '"₩"#,##0;[Red]\\-"₩"#,##0'

# 날짜별 색상 팔레트 (같은 날 복수 결제 시 하이라이트)
DATE_COLORS = [
    PatternFill(fgColor="FFFFFF00", fill_type="solid"),  # 노랑
    PatternFill(fgColor="FFC6EFCE", fill_type="solid"),  # 연두
    PatternFill(fgColor="FFBDD7EE", fill_type="solid"),  # 하늘
    PatternFill(fgColor="FFF8CBAD", fill_type="solid"),  # 연분홍
    PatternFill(fgColor="FFE2EFDA", fill_type="solid"),  # 연한 초록
]

# 확인 필요 행 강조 (사람이 반드시 손대야 하는 행)
REVIEW_FILL = PatternFill(fgColor="FFFFC7CE", fill_type="solid")  # 연한 빨강
# 초과금액이 발생한 셀 강조
EXCESS_FILL = PatternFill(fgColor="FFFCE4D6", fill_type="solid")  # 연한 주황

# 칼럼 너비
# 결제금액 / 인정금액 / 초과금액을 각각 따로 두는 것이 이 양식의 핵심이다.
# 한도를 넘겼을 때 원 결제금액을 덮어쓰지 않고 초과분만 분리해 보여준다.
COL_WIDTHS = {
    "A": 4.5,     # NO.
    "B": 10.625,  # 사용날짜
    "C": 25.25,   # 사용처
    "D": 10.625,  # 항목 구분
    "E": 15.5,    # 사용내역
    "F": 12.5,    # 결제금액 (영수증 원금액)
    "G": 12.5,    # 인정금액 (한도 내 = 실제 청구액)
    "H": 12.5,    # 초과금액 (한도 초과분)
    "I": 26.0,    # 참석자
    "J": 34.0,    # 비고
}

LAST_COL = 10  # J
LAST_COL_LETTER = "J"
# 항목 구분 데이터검증 목록을 숨겨둘 칼럼 (데이터 영역 밖)
VALIDATION_COL = 13  # M
VALIDATION_COL_LETTER = "M"

# 데이터 검증 항목 목록
CATEGORY_LIST = ["식대", "여비교통비", "회식비", "접대비", "소모품비"]

# Sheet 2 영수증 갤러리 설정
RECEIPT_COLS_PER_IMAGE = 5  # 이미지당 칼럼 수
RECEIPT_GAP_COLS = 1  # 이미지 블록 간 좌우 간격 칼럼
RECEIPT_GAP_ROWS = 2  # 이미지 블록 간 상하 간격 행
RECEIPT_ROWS_PER_IMAGE = 35  # 이미지당 행 수 (라벨 1행 + 이미지 34행)
IMAGES_PER_ROW = 5  # 한 줄에 5개 이미지
RECEIPT_MAX_WIDTH = 380  # 이미지 최대 너비 (px)
RECEIPT_MAX_HEIGHT = 700  # 이미지 최대 높이 (px)


def create_expense_excel(
    result: ProcessResult,
    user_name: str,
    department: str,
    position: str,
    submit_date: date | None = None,
    receipt_images: list[tuple[str, str]] | None = None,
    output_path: str | None = None,
    limit_summary: list[str] | None = None,
) -> str:
    """회사 양식에 맞는 엑셀 파일을 생성한다.

    Args:
        result: 판정 결과 (행 목록 + 적용된 한도)
        user_name: 성명
        department: 부서
        position: 직책
        submit_date: 제출일 (None이면 오늘)
        receipt_images: (파일경로, 라벨) 튜플 리스트 (Sheet 2용)
        output_path: 출력 파일 경로 (None이면 자동 생성)
        limit_summary: 적용된 한도와 출처를 설명하는 문장들 (시트 하단에 기록)

    Returns:
        생성된 파일 경로
    """
    if submit_date is None:
        submit_date = date.today()

    rows = result.rows
    wb = Workbook()

    # Sheet 1: 개인경비 사용내역서
    ws = wb.active
    ws.title = "개인경비 사용내역서"
    _build_expense_sheet(
        ws, result, user_name, department, position, submit_date, limit_summary or []
    )

    # Sheet 2: 영수증 (이미지가 있는 경우)
    if receipt_images:
        ws2 = wb.create_sheet("영수증")
        _build_receipt_sheet(ws2, receipt_images)

    # 파일 저장
    if output_path is None:
        # 데이터 기준 월 자동 결정
        if rows and rows[0].date:
            ref_date = rows[0].date
        else:
            ref_date = submit_date
        filename = f"{ref_date.year}{ref_date.month:02d}_개인경비_사용내역서_{user_name}.xlsx"
        output_path = os.path.join(os.path.expanduser("~"), "Downloads", filename)

    wb.save(output_path)
    return output_path


def _build_expense_sheet(
    ws,
    result: ProcessResult,
    user_name: str,
    department: str,
    position: str,
    submit_date: date,
    limit_summary: list[str],
):
    """Sheet 1: 개인경비 사용내역서를 구성한다."""
    rows = result.rows

    # 칼럼 너비 설정
    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[col].width = width

    # 숨김 칼럼 (항목 구분 데이터 검증 목록)
    ws.column_dimensions[VALIDATION_COL_LETTER].width = 13.25
    ws.column_dimensions[VALIDATION_COL_LETTER].hidden = True

    # === Row 1: 제목 ===
    ws.merge_cells(f"A1:{LAST_COL_LETTER}1")
    ws.row_dimensions[1].height = 26.25
    cell = ws["A1"]
    cell.value = "개인경비 사용내역서"
    cell.font = FONT_TITLE
    cell.alignment = ALIGN_CENTER

    # === Row 3: 성명 / 부서 ===
    ws["B3"].value = "성 명   :"
    ws["B3"].font = FONT_LABEL
    ws["B3"].alignment = ALIGN_RIGHT
    ws["C3"].value = user_name
    ws["C3"].font = FONT_DATA

    ws["I3"].value = "부 서(Dep't)  :"
    ws["I3"].font = FONT_LABEL
    ws["I3"].alignment = ALIGN_RIGHT
    ws["J3"].value = department
    ws["J3"].font = FONT_DATA

    # === Row 4: 직책 / 제출일 ===
    ws["B4"].value = "직 책   :"
    ws["B4"].font = FONT_LABEL
    ws["B4"].alignment = ALIGN_RIGHT
    ws["C4"].value = position
    ws["C4"].font = FONT_DATA

    ws["I4"].value = "제출일(Date) :"
    ws["I4"].font = FONT_LABEL
    ws["I4"].alignment = ALIGN_RIGHT
    ws["J4"].value = submit_date
    ws["J4"].font = FONT_DATA
    ws["J4"].number_format = "YYYY.MM.DD"

    # === Row 5: 헤더 ===
    ws.row_dimensions[5].height = 30
    headers = [
        ("A5", "NO."),
        ("B5", "사용날짜"),
        ("C5", "사용처\n(전표상 상호)"),
        ("D5", "항목 구분"),
        ("E5", "사용내역"),
        ("F5", "결제금액"),
        ("G5", "인정금액\n(한도 내)"),
        ("H5", "초과금액\n(한도 초과)"),
        ("I5", "참석자(대상)"),
        ("J5", "비고"),
    ]
    for cell_ref, text in headers:
        cell = ws[cell_ref]
        cell.value = text
        cell.font = FONT_HEADER_SMALL if "\n" in text else FONT_HEADER
        cell.alignment = ALIGN_CENTER_WRAP if "\n" in text else ALIGN_CENTER
        cell.border = THIN_BORDER

    # 데이터 검증 목록 (숨김 칼럼)
    for i, cat in enumerate(CATEGORY_LIST):
        ws.cell(row=5 + i, column=VALIDATION_COL).value = cat

    first_v = f"${VALIDATION_COL_LETTER}$5"
    last_v = f"${VALIDATION_COL_LETTER}${4 + len(CATEGORY_LIST)}"
    dv = DataValidation(type="list", formula1=f"={first_v}:{last_v}", allow_blank=True)
    dv.error = "목록에서 항목을 선택하세요"
    dv.errorTitle = "항목 선택"
    ws.add_data_validation(dv)

    # === 날짜별 색상 매핑 ===
    # 같은 날짜에 2건 이상인 경우만 색칠 (= 한도를 나눠 쓴 날이라는 신호)
    from collections import Counter

    date_counts = Counter(r.date for r in rows if r.date is not None and not r.needs_review)
    multi_dates = {d for d, cnt in date_counts.items() if cnt >= 2}
    date_color_map: dict = {}
    color_idx = 0
    for r in rows:
        if r.date in multi_dates and r.date not in date_color_map:
            date_color_map[r.date] = DATE_COLORS[color_idx % len(DATE_COLORS)]
            color_idx += 1

    # === 데이터 행 ===
    max_data_rows = max(len(rows), 35)  # 최소 35행 (여백)
    for i in range(max_data_rows):
        row_num = 6 + i
        ws.row_dimensions[row_num].height = 21

        expense = rows[i] if i < len(rows) else None

        if expense is not None:
            date_str = expense.date.strftime("%y.%m.%d") if expense.date else ""

            ws.cell(row=row_num, column=1, value=expense.seq_num)
            ws.cell(row=row_num, column=2, value=date_str)
            ws.cell(row=row_num, column=3, value=expense.merchant)
            ws.cell(row=row_num, column=4, value=expense.category)
            ws.cell(row=row_num, column=5, value=expense.description)
            # 금액을 판독하지 못한 건은 0 이 아니라 빈 칸으로 둔다.
            # 0 을 넣으면 "0원짜리 결제"로 읽혀 사람이 오류를 놓친다.
            ws.cell(row=row_num, column=6, value=expense.amount)
            if expense.needs_review:
                ws.cell(row=row_num, column=7, value=None)
                ws.cell(row=row_num, column=8, value=None)
            else:
                ws.cell(row=row_num, column=7, value=expense.claimed)
                ws.cell(row=row_num, column=8, value=expense.excess if expense.excess else None)
            ws.cell(row=row_num, column=9, value=expense.attendees)
            ws.cell(row=row_num, column=10, value=expense.note)
        else:
            ws.cell(row=row_num, column=1, value=i + 1)

        # 스타일
        for col in range(1, LAST_COL + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = FONT_DATA
            cell.border = THIN_BORDER
            if col == 3:  # 사용처
                cell.alignment = ALIGN_CENTER_SHRINK
            elif col in (9, 10):  # 참석자, 비고
                cell.alignment = ALIGN_LEFT
            elif col in (6, 7, 8):  # 금액 3종
                cell.alignment = ALIGN_CENTER
                cell.number_format = WON_FORMAT
            else:
                cell.alignment = ALIGN_CENTER

        if expense is None:
            dv.add(ws.cell(row=row_num, column=4))
            continue

        if expense.needs_review:
            # 확인 필요 행: 행 전체를 빨갛게 — 이 행은 그대로 제출하면 안 된다.
            for col in range(1, LAST_COL + 1):
                ws.cell(row=row_num, column=col).fill = REVIEW_FILL
        else:
            # 같은 날 복수 결제 하이라이트 (날짜 + 금액 블록)
            if expense.date in date_color_map:
                fill = date_color_map[expense.date]
                for col in (2, 6, 7, 8):
                    ws.cell(row=row_num, column=col).fill = fill
            # 초과분이 있는 셀은 따로 강조
            if expense.excess:
                ws.cell(row=row_num, column=8).fill = EXCESS_FILL

        dv.add(ws.cell(row=row_num, column=4))

    last_data_row = 5 + max_data_rows

    # === 합계 행 ===
    total_row = 6 + max_data_rows
    ws.row_dimensions[total_row].height = 22
    ws.merge_cells(f"A{total_row}:E{total_row}")
    ws.merge_cells(f"I{total_row}:J{total_row}")

    total_cell = ws.cell(row=total_row, column=1)
    total_cell.value = "합계"
    total_cell.font = FONT_HEADER
    total_cell.alignment = ALIGN_CENTER
    for col in range(1, 6):
        ws.cell(row=total_row, column=col).border = THIN_BORDER

    # F/G/H 각각 합계. 확인필요 행은 G/H 가 비어 있으므로 자연히 제외된다.
    for col in (6, 7, 8):
        letter = get_column_letter(col)
        cell = ws.cell(row=total_row, column=col)
        cell.value = f"=SUM({letter}6:{letter}{last_data_row})"
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
        cell.number_format = WON_FORMAT
        cell.border = THIN_BORDER

    claim_note = ws.cell(row=total_row, column=9)
    claim_note.value = "← 인정금액이 실제 청구액입니다"
    claim_note.font = FONT_NOTE
    claim_note.alignment = ALIGN_LEFT
    for col in range(9, LAST_COL + 1):
        ws.cell(row=total_row, column=col).border = THIN_BORDER

    # === 확인 필요 안내 (해당 건이 있을 때만) ===
    cursor = total_row + 2
    review_rows = result.review_rows
    if review_rows:
        ws.merge_cells(f"A{cursor}:{LAST_COL_LETTER}{cursor}")
        warn = ws.cell(row=cursor, column=1)
        warn.value = (
            f"⚠ 확인 필요 {len(review_rows)}건 — 붉은색 행은 영수증에서 날짜/금액을 "
            f"읽지 못한 건입니다. 원본을 확인해 직접 입력한 뒤 제출하세요 "
            f"(합계에 반영되어 있지 않습니다)."
        )
        warn.font = FONT_NOTE_BOLD
        warn.alignment = ALIGN_LEFT
        warn.fill = REVIEW_FILL
        cursor += 2

    # === 적용된 한도와 출처 (감사 추적용) ===
    if limit_summary:
        ws.cell(row=cursor, column=1).value = "적용된 일 한도"
        ws.cell(row=cursor, column=1).font = FONT_LABEL
        cursor += 1
        for line in limit_summary:
            ws.cell(row=cursor, column=1).value = line
            ws.cell(row=cursor, column=1).font = FONT_NOTE
            cursor += 1
        cursor += 1

    # === 하단 안내 문구 ===
    ws.cell(row=cursor, column=1).value = "* 사용일 순, 개별건으로 정리하여 입력"
    ws.cell(row=cursor, column=1).font = FONT_NOTE
    ws.cell(row=cursor + 1, column=1).value = "* 항목은 박스우측상단을 클릭한 후 항목중에서 선택"
    ws.cell(row=cursor + 1, column=1).font = FONT_NOTE
    ws.cell(row=cursor + 2, column=1).value = "* 참석자가 여러명일 경우 전원의 이름 및 소속 기입"
    ws.cell(row=cursor + 2, column=1).font = FONT_NOTE_BOLD

    # 페이지 설정
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.scale = 80
    ws.freeze_panes = "A6"


def _build_receipt_sheet(
    ws,
    receipt_images: list[tuple[str, str]],
):
    """Sheet 2: 영수증 이미지 갤러리를 구성한다."""
    ws.sheet_view.zoomScale = 70

    # 이미지 블록 간격 포함 칼럼 폭 계산
    block_width = RECEIPT_COLS_PER_IMAGE + RECEIPT_GAP_COLS  # 5 + 1 = 6

    for idx, (img_path, label) in enumerate(receipt_images):
        # 그리드 위치 계산 (간격 칼럼 포함)
        grid_row = idx // IMAGES_PER_ROW
        grid_col = idx % IMAGES_PER_ROW

        start_col = grid_col * block_width + 1  # 1-based
        start_row = grid_row * (RECEIPT_ROWS_PER_IMAGE + RECEIPT_GAP_ROWS) + 1  # 1-based

        # 라벨 셀 (병합)
        end_col_letter = get_column_letter(start_col + RECEIPT_COLS_PER_IMAGE - 1)
        start_col_letter = get_column_letter(start_col)
        label_range = f"{start_col_letter}{start_row}:{end_col_letter}{start_row}"
        ws.merge_cells(label_range)

        label_cell = ws.cell(row=start_row, column=start_col)
        label_cell.value = label
        label_cell.font = FONT_DATA
        label_cell.alignment = ALIGN_CENTER
        label_cell.border = Border(top=Side(style="thin"), left=Side(style="thin"))

        # 이미지 영역 병합
        img_start_row = start_row + 1
        img_end_row = start_row + RECEIPT_ROWS_PER_IMAGE - 1
        img_range = f"{start_col_letter}{img_start_row}:{end_col_letter}{img_end_row}"
        ws.merge_cells(img_range)

        # 이미지 삽입 (원본 비율 유지)
        if os.path.exists(img_path):
            try:
                img = XlImage(img_path)
                orig_w = img.width
                orig_h = img.height
                if orig_w > 0 and orig_h > 0:
                    ratio = min(RECEIPT_MAX_WIDTH / orig_w, RECEIPT_MAX_HEIGHT / orig_h)
                    img.width = int(orig_w * ratio)
                    img.height = int(orig_h * ratio)
                anchor = f"{start_col_letter}{img_start_row}"
                ws.add_image(img, anchor)
            except Exception:
                ws.cell(row=img_start_row, column=start_col).value = f"[이미지: {os.path.basename(img_path)}]"
