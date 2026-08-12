"""CLI 레벨 테스트 — 실제로 스크립트를 돌려 엑셀이 나오고 한도/출처가 출력되는지 확인."""

import json
import os
import subprocess
import sys
import tempfile

sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "scripts"))

from test_support import run_test_classes  # noqa: E402

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SCRIPT = os.path.join(ROOT, "scripts", "generate_excel.py")

RECEIPTS = [
    {"date": "2026-06-20", "merchant": "행복분식", "amount": 20000, "image_path": ""},
    {"date": "2026-07-20", "merchant": "카페 하루", "amount": 20000, "image_path": ""},
    {"date": "2026-07-21", "merchant": "미래식당", "amount": 8000, "image_path": ""},
    {"date": None, "merchant": "", "amount": None, "image_path": ""},
]

CONFIG = {
    "daily_limit": {
        "default": 15000,
        "history": [
            {"from": "2025-01-01", "amount": 15000, "note": "구규정"},
            {"from": "2026-07-01", "amount": 20000, "note": "신규정"},
        ],
    }
}


def _tmp_json(payload, suffix=".json") -> str:
    fd, path = tempfile.mkstemp(suffix=suffix)
    os.close(fd)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False)
    return path


def _run(extra_args=None, receipts=None, config=None):
    """generate_excel.py 를 서브프로세스로 실행한다.

    uv 가 부모 프로세스의 VIRTUAL_ENV 를 상속해 엉뚱한 환경을 잡는 것을 막기 위해
    관련 환경변수를 비운다.
    """
    receipts_path = _tmp_json(receipts if receipts is not None else RECEIPTS)
    config_path = _tmp_json(config if config is not None else CONFIG)
    out_path = os.path.join(tempfile.mkdtemp(), "결과.xlsx")

    env = dict(os.environ)
    env.pop("VIRTUAL_ENV", None)
    env["PYTHONIOENCODING"] = "utf-8"
    # 인적사항은 저장된 설정 대신 인자로 넘겨 사용자 홈을 건드리지 않는다.
    env["SIKDAE_CONFIG_PATH"] = os.path.join(tempfile.mkdtemp(), "nonexistent.json")

    cmd = [
        sys.executable, SCRIPT, receipts_path,
        "--output", out_path,
        "--config", config_path,
        "--name", "홍길동", "--position", "매니저", "--department", "테스트팀",
    ] + (extra_args or [])
    proc = subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8", env=env, cwd=ROOT)
    return proc, out_path


class TestLimitSourceOutput:
    """실행 결과에 적용된 한도와 출처가 반드시 나와야 한다."""

    def test_config_history_source_is_reported(self):
        proc, _ = _run()
        assert proc.returncode == 0, proc.stderr
        assert "config.json 구간" in proc.stdout, proc.stdout

    def test_both_brackets_reported_when_receipts_span_rule_change(self):
        proc, _ = _run()
        assert "15,000원" in proc.stdout, proc.stdout
        assert "20,000원" in proc.stdout, proc.stdout

    def test_arg_override_source_is_reported(self):
        proc, _ = _run(["--daily-limit", "18000"])
        assert proc.returncode == 0, proc.stderr
        assert "18,000원" in proc.stdout, proc.stdout
        assert "실행 인자" in proc.stdout, proc.stdout

    def test_chat_override_source_is_reported(self):
        proc, _ = _run(["--daily-limit-from-chat", "20000"])
        assert proc.returncode == 0, proc.stderr
        assert "대화 지시" in proc.stdout, proc.stdout

    def test_chat_beats_arg_in_cli(self):
        proc, _ = _run(["--daily-limit", "11000", "--daily-limit-from-chat", "20000"])
        assert "대화 지시" in proc.stdout, proc.stdout
        assert "20,000원" in proc.stdout, proc.stdout
        assert "실행 인자" not in proc.stdout, proc.stdout

    def test_missing_config_reports_builtin_default(self):
        receipts_path = _tmp_json(RECEIPTS)
        env = dict(os.environ)
        env.pop("VIRTUAL_ENV", None)
        env["PYTHONIOENCODING"] = "utf-8"
        env["SIKDAE_CONFIG_PATH"] = os.path.join(tempfile.mkdtemp(), "none.json")
        out = os.path.join(tempfile.mkdtemp(), "o.xlsx")
        proc = subprocess.run(
            [sys.executable, SCRIPT, receipts_path, "--output", out,
             "--config", os.path.join(tempfile.mkdtemp(), "missing.json"), "--name", "홍길동"],
            capture_output=True, text=True, encoding="utf-8", env=env, cwd=ROOT,
        )
        assert proc.returncode == 0, proc.stderr
        assert "내장 기본값" in proc.stdout, proc.stdout


class TestTotalsOutput:
    def test_totals_are_printed(self):
        proc, _ = _run()
        # 2026-06-20 20000 → 한도 15000 (초과 5000)
        # 2026-07-20 20000 → 한도 20000 (초과 0)
        # 2026-07-21  8000 → 한도 20000 (초과 0)
        assert "인정금액 합계 : 43,000원" in proc.stdout, proc.stdout
        assert "초과금액 합계 : 5,000원" in proc.stdout, proc.stdout

    def test_review_items_are_listed(self):
        proc, _ = _run()
        assert "확인 필요     : 1건" in proc.stdout, proc.stdout
        assert "사용일자 미판독" in proc.stdout, proc.stdout

    def test_arg_limit_changes_totals(self):
        proc, _ = _run(["--daily-limit", "10000"])
        # 세 건 모두 10000 한도 → 10000 + 10000 + 8000 = 28000
        assert "인정금액 합계 : 28,000원" in proc.stdout, proc.stdout


class TestExcelOutput:
    def test_excel_file_is_created(self):
        proc, out_path = _run()
        assert proc.returncode == 0, proc.stderr
        assert os.path.exists(out_path), f"엑셀이 생성되지 않음: {proc.stdout}\n{proc.stderr}"
        assert os.path.getsize(out_path) > 3000, os.path.getsize(out_path)

    def test_excel_has_excess_column_and_values(self):
        from openpyxl import load_workbook

        proc, out_path = _run()
        assert proc.returncode == 0, proc.stderr
        ws = load_workbook(out_path).active

        headers = [ws.cell(row=5, column=c).value for c in range(1, 11)]
        assert any("결제금액" in str(h) for h in headers), headers
        assert any("인정금액" in str(h) for h in headers), headers
        assert any("초과금액" in str(h) for h in headers), headers

        # 첫 데이터 행 = 2026-06-20, 20000원, 한도 15000
        assert ws.cell(row=6, column=6).value == 20000, ws.cell(row=6, column=6).value
        assert ws.cell(row=6, column=7).value == 15000, ws.cell(row=6, column=7).value
        assert ws.cell(row=6, column=8).value == 5000, ws.cell(row=6, column=8).value

    def test_review_row_amount_left_blank_not_zero(self):
        from openpyxl import load_workbook

        proc, out_path = _run()
        ws = load_workbook(out_path).active
        # 확인필요 건은 4번째(마지막) 데이터 행
        assert ws.cell(row=9, column=6).value is None, ws.cell(row=9, column=6).value
        assert ws.cell(row=9, column=7).value is None, ws.cell(row=9, column=7).value
        assert "확인 필요" in str(ws.cell(row=9, column=10).value), ws.cell(row=9, column=10).value

    def test_excel_records_applied_limit(self):
        from openpyxl import load_workbook

        proc, out_path = _run(["--daily-limit-from-chat", "20000"])
        ws = load_workbook(out_path).active
        text = " ".join(
            str(ws.cell(row=r, column=1).value or "")
            for r in range(1, ws.max_row + 1)
        )
        assert "적용된 일 한도" in text, text
        assert "대화 지시" in text, text


class TestErrorHandling:
    def test_missing_name_errors_out(self):
        receipts_path = _tmp_json(RECEIPTS)
        env = dict(os.environ)
        env.pop("VIRTUAL_ENV", None)
        env["PYTHONIOENCODING"] = "utf-8"
        env["SIKDAE_CONFIG_PATH"] = os.path.join(tempfile.mkdtemp(), "none.json")
        proc = subprocess.run(
            [sys.executable, SCRIPT, receipts_path],
            capture_output=True, text=True, encoding="utf-8", env=env, cwd=ROOT,
        )
        assert proc.returncode == 1, proc.stdout
        assert "성명" in proc.stderr, proc.stderr

    def test_negative_limit_rejected(self):
        proc, _ = _run(["--daily-limit", "-5000"])
        assert proc.returncode == 1, proc.stdout
        assert "0보다 커야" in proc.stderr, proc.stderr

    def test_empty_receipts_rejected(self):
        proc, _ = _run(receipts=[])
        assert proc.returncode == 1, proc.stdout
        assert "비어" in proc.stderr, proc.stderr

    def test_malformed_date_becomes_review_not_guess(self):
        proc, _ = _run(receipts=[{"date": "2026-13-45", "merchant": "A", "amount": 9000}])
        assert proc.returncode == 0, proc.stderr
        assert "확인 필요     : 1건" in proc.stdout, proc.stdout


if __name__ == "__main__":
    ok = run_test_classes(
        TestLimitSourceOutput, TestTotalsOutput, TestExcelOutput, TestErrorHandling
    )
    sys.exit(0 if ok else 1)
